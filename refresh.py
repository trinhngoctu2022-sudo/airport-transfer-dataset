"""
refresh.py — Chay sau moi lan sua file Excel.

Lam 3 viec:
  1. Doc metravel_transport_dataset.xlsx
  2. Xuat ra transport_prices.csv
  3. Nap vao metravel.db va bao cao chat luong du lieu

Cach chay (mo Terminal, cd vao thu muc chua file nay):
    python3 refresh.py

Lan dau can cai thu vien:
    pip3 install openpyxl
"""

import csv
import os
import sqlite3
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Chua cai openpyxl. Chay: pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "metravel_transport_dataset.xlsx")
CSV_OUT = os.path.join(HERE, "transport_prices.csv")
DB = os.path.join(HERE, "metravel.db")

NUMERIC = {
    "price_min_local", "price_max_local", "price_min_vnd", "price_max_vnd",
    "duration_min", "duration_max", "comfort_1_5", "language_difficulty_1_5",
}

# ---------------------------------------------------------------- 1. Doc Excel
if not os.path.exists(XLSX):
    sys.exit(f"Khong tim thay {XLSX}")

wb = load_workbook(XLSX, data_only=True)
ws = wb["transport_data"]
header = [c.value for c in ws[1]]

rows = []
for r in range(3, ws.max_row + 1):          # bo dong 2 (EXAMPLE)
    vals = [ws.cell(r, i).value for i in range(1, len(header) + 1)]
    if any(v not in (None, "") for v in vals[8:]):   # chi lay dong da dien
        rows.append(vals)

if not rows:
    sys.exit("Chua co dong nao duoc dien. Dien du lieu vao Excel truoc.")

# Canh bao neu cot VND rong -> quen bam tinh lai cong thuc
vnd_empty = sum(1 for v in rows if v[10] in (None, ""))
if vnd_empty:
    print(f"  CANH BAO: {vnd_empty} dong co cot VND rong.")
    print("   Mo Excel, bam Cmd+= de tinh lai cong thuc, luu lai, roi chay lai script nay.\n")

# ---------------------------------------------------------------- 2. Xuat CSV
with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(header)
    w.writerows(rows)
print(f"Da xuat {len(rows)} dong -> transport_prices.csv")

# ---------------------------------------------------------------- 3. Nap SQLite
con = sqlite3.connect(DB)
cur = con.cursor()
cur.execute("DROP TABLE IF EXISTS routes")
cur.execute(
    "CREATE TABLE routes ("
    + ",".join(f'"{c}" ' + ("REAL" if c in NUMERIC else "TEXT") for c in header)
    + ")"
)

for row in rows:
    clean = []
    for col, val in zip(header, row):
        if val in (None, "", "None"):
            clean.append(None)
        elif col in NUMERIC:
            try:
                clean.append(float(val))
            except (TypeError, ValueError):
                clean.append(None)
        else:
            clean.append(str(val))
    cur.execute(
        "INSERT INTO routes VALUES (" + ",".join("?" * len(header)) + ")", clean
    )
con.commit()
print(f"Da nap vao metravel.db (bang 'routes')")

# ---------------------------------------------------------------- 4. Kiem tra
print("\n--- KIEM TRA CHAT LUONG ---")
checks = [
    ("Thieu gia",              "price_min_local IS NULL"),
    ("Thieu comfort",          "comfort_1_5 IS NULL"),
    ("Thieu do kho ngon ngu",  "language_difficulty_1_5 IS NULL"),
    ("Thieu last_updated",     "last_updated IS NULL"),
    ("Thieu nguon",            "source_url IS NULL"),
    ("luggage_ok sai chuan",   "luggage_ok IS NOT NULL AND luggage_ok NOT IN ('yes','tight','no')"),
    ("Gia min > max",          "price_min_local > price_max_local"),
    ("Chua xac minh",          "scam_risk_note LIKE '%CHUA XAC MINH%'"),
    ("Thieu price_basis",      "price_basis IS NULL"),
    ("price_basis sai chuan",  "price_basis IS NOT NULL AND price_basis NOT IN ('per_person','per_vehicle')"),
]

problems = 0
for label, cond in checks:
    n = cur.execute(f"SELECT COUNT(*) FROM routes WHERE {cond}").fetchone()[0]
    problems += n
    mark = "OK  " if n == 0 else "FIX "
    print(f"  [{mark}] {label}: {n}")

print()
if problems == 0:
    print("Du lieu sach. San sang chay analysis.sql.")
else:
    print(f"Con {problems} van de can sua trong Excel truoc khi phan tich.")

con.close()
